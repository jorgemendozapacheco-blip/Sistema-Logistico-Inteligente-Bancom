import datetime
import openpyxl
from django.db import transaction
from .models import Item

# --- Cabeceras y mapeo ---
EXCEL_HEADERS = [
    'ID del CI', 'Número de etiqueta', 'Número de serie', 'Estado',
    'Motivo del estado', 'Nombre del producto', 'Modelo/versión',
    'Fabricante', 'Nombre del proveedor', 'Localidad', 'Oficina', 'Piso',
    'Fecha de adquisición', 'ProcessorType', 'DiskCapacityGB',
    'OSArchitecture', 'OSName', 'OSType', 'IPAddress1',
    'TotalPhysicalMemory', 'Nombre Completo', 'Departamento [PER]',
    'Zona',   # <-- Nuevo campo importante
]

FIELD_MAP = {
    'ID del CI':            'codigo',
    'Número de etiqueta':   'etiqueta',
    'Número de serie':      'numerodeserie',
    'Estado':               'estado',
    'Motivo del estado':    'motivo_estado',
    'Nombre del producto':  'nombre',
    'Modelo/versión':       'modelo',
    'Fabricante':           'marca',
    'Nombre del proveedor': 'proveedor',
    'Localidad':            'localidad',
    'Oficina':              'oficina',
    'Piso':                 'piso',
    'Fecha de adquisición': 'fecha_adquisicion',
    'ProcessorType':        'cpu_type',
    'DiskCapacityGB':       'disk_gb',
    'OSArchitecture':       'os_arch',
    'OSName':               'os_name',
    'OSType':               'os_type',
    'IPAddress1':           'ip_address',
    'TotalPhysicalMemory':  'ram_gb',
    'Nombre Completo':      'empleado',
    'Departamento [PER]':   'dept_per',
    'Zona':                 'zona',
}

NUMERIC_FIELDS = {'disk_gb', 'ram_gb'}

def _to_int(val):
    try:
        return int(float(val))
    except Exception:
        return 0

def _to_date(val):
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                return datetime.datetime.strptime(val.strip(), fmt).date()
            except Exception:
                continue
    return None

def importar_excel(path, batch_size=2000):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {
        FIELD_MAP[col.strip()]: idx
        for idx, col in enumerate(header_row)
        if col and col.strip() in EXCEL_HEADERS
    }

    batch = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        data = {}
        for campo, idx in col_index.items():
            val = row[idx]
            if campo in NUMERIC_FIELDS:
                data[campo] = _to_int(val)
            elif campo == 'fecha_adquisicion':
                data[campo] = _to_date(val)
            else:
                data[campo] = val if val is not None else ''

        if not data.get('codigo'):
            continue

        batch.append(Item(**data))

        # Carga en lotes (más eficiente en lotes grandes)
        if len(batch) >= batch_size:
            with transaction.atomic():
                Item.objects.bulk_create(batch, ignore_conflicts=True)
            batch.clear()  # más eficiente que batch = []

    if batch:
        with transaction.atomic():
            Item.objects.bulk_create(batch, ignore_conflicts=True)
        batch.clear()